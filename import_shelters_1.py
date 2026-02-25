import os
import re
import uuid
import argparse
from datetime import datetime
from zoneinfo import ZoneInfo

from dotenv import load_dotenv
import psycopg
from psycopg.rows import dict_row
from openpyxl import load_workbook

load_dotenv()

# ===== Table names (ปรับให้ตรง DB ของคุณ) =====
T_SHELTER = "Shelter"
T_DISASTERS = "disasters"
T_FACILITY = "ShelterFacility"
T_DAILYSTATS = "ShelterDailyStats"
T_RESIDENT = "Resident"

TZ = ZoneInfo("Asia/Bangkok")


def qname(name: str) -> str:
    return f'"{name}"' if any(c.isupper() for c in name) else name


def get_conn():
    host = os.getenv("PGHOST", "127.0.0.1")
    port = int(os.getenv("PGPORT", "5432"))
    dbname = os.getenv("PGDATABASE")
    user = os.getenv("PGUSER")
    password = os.getenv("PGPASSWORD")
    sslmode = os.getenv("PGSSL")

    if not all([dbname, user, password]):
        raise ValueError("Missing env: PGDATABASE / PGUSER / PGPASSWORD")

    return psycopg.connect(
        host=host,
        port=port,
        dbname=dbname,
        user=user,
        password=password,
        sslmode=sslmode,
        connect_timeout=15,
        autocommit=False,
        row_factory=dict_row,
    )


# ------------------------------
# helpers
# ------------------------------
def norm_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None


def is_has(v) -> bool:
    s = norm_text(v)
    if not s:
        return False

    s = s.strip()

    # เคสปฏิเสธ (กัน false positive)
    if "ไม่มี" in s or "ไม่รองรับ" in s or "ไม่รับรอง" in s or s.startswith("ไม่"):
        return False

    # เคสยืนยัน
    if s == "มี":
        return True
    if "มี" in s:
        return True
    if "รองรับ" in s:
        return True
    if "รับรอง" in s:
        return True

    return False


def to_int(v, default=None):
    s = norm_text(v)
    if s is None:
        return default
    try:
        return int(float(s))
    except Exception:
        return default


def to_float(v, default=None):
    s = norm_text(v)
    if s is None:
        return default
    try:
        return float(s)
    except Exception:
        return default


def clean_phone(v):
    s = norm_text(v)
    if not s:
        return None
    s = s.replace(" ", "").replace("-", "")
    s = re.sub(r"[^0-9+]", "", s)
    return s or None


def map_shelter_type(v):
    s = norm_text(v)
    if not s:
        return "SYSTEM"

    # ถ้าเจอคำว่า ศูนย์พักพิงพระราชทาน ให้เป็น ROYAL
    if "ศูนย์พักพิงพระราชทาน" in s:
        return "ROYAL"

    # นอกนั้นทั้งหมดเป็น SYSTEM
    return "SYSTEM"


# ------------------------------
# Excel reader
# ------------------------------
def read_excel_rows(excel_path, sheet_name=None, start_row=2):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = []
    for r in range(start_row, ws.max_row + 1):
        # ✅ กันแถว "ว่างจริง" (ไม่งั้นจะ insert แถวท้ายไฟล์เป็นร้อย)
        probe = [
            ws[f"A{r}"].value, ws[f"B{r}"].value, ws[f"C{r}"].value, ws[f"D{r}"].value,
            ws[f"E{r}"].value, ws[f"F{r}"].value, ws[f"G{r}"].value,
            ws[f"H{r}"].value, ws[f"I{r}"].value, ws[f"J{r}"].value, ws[f"K{r}"].value,
            ws[f"M{r}"].value, ws[f"N{r}"].value, ws[f"Q{r}"].value, ws[f"AC{r}"].value,
            ws[f"S{r}"].value, ws[f"T{r}"].value, ws[f"U{r}"].value, ws[f"V{r}"].value,
            ws[f"W{r}"].value, ws[f"X{r}"].value, ws[f"Y{r}"].value, ws[f"Z{r}"].value,
            ws[f"AA{r}"].value, ws[f"AB{r}"].value,
        ]
        if all(norm_text(v) is None for v in probe):
            continue

        # ✅ ไม่สน name ว่าง: ถ้าว่างให้เป็น ""
        name = norm_text(ws[f"A{r}"].value)
        if name is None:
            name = ""

        row = {
            "name": name,
            "shelterType": map_shelter_type(ws[f"B{r}"].value),

            # ✅ เปลี่ยนจาก village -> addressLine (เดิม column C)
            "addressLine": norm_text(ws[f"C{r}"].value),

            "province": norm_text(ws[f"D{r}"].value),
            "district": norm_text(ws[f"E{r}"].value),
            "subdistrict": norm_text(ws[f"F{r}"].value),
            "postcode": norm_text(ws[f"G{r}"].value),

            "coordinatorName": norm_text(ws[f"H{r}"].value),
            "coordinatorPhone": clean_phone(ws[f"I{r}"].value),

            "latitude": to_float(ws[f"J{r}"].value),
            "longitude": to_float(ws[f"K{r}"].value),

            "capacity": to_int(ws[f"M{r}"].value, default=None),
            "restroomCount": to_int(ws[f"N{r}"].value, default=0),
            "parkingCount": to_int(ws[f"Q{r}"].value, default=0),

            "responsibleAgency": norm_text(ws[f"AC{r}"].value),

            # facilities flags: S,T,U,V,AA,AB
            "fac_s": is_has(ws[f"S{r}"].value),     # ELECTRICITY
            "fac_t": is_has(ws[f"T{r}"].value),     # TENT
            "fac_u": is_has(ws[f"U{r}"].value),     # PET
            "fac_v": is_has(ws[f"V{r}"].value),     # KITCHEN
            "fac_aa": is_has(ws[f"AA{r}"].value),   # INTERNET
            "fac_ab": is_has(ws[f"AB{r}"].value),   # SECURITY

            # disasters flags: W,X,Y,Z
            "dis_w": is_has(ws[f"W{r}"].value),     # FLOOD
            "dis_x": is_has(ws[f"X{r}"].value),     # STORM
            "dis_y": is_has(ws[f"Y{r}"].value),     # FIRE
            "dis_z": is_has(ws[f"Z{r}"].value),     # EARTHQUAKE
        }
        rows.append(row)

    return rows


def build_disaster_types(row):
    out = []
    if row.get("dis_w"):
        out.append("FLOOD")
    if row.get("dis_x"):
        out.append("STORM")
    if row.get("dis_y"):
        out.append("FIRE")
    if row.get("dis_z"):
        out.append("EARTHQUAKE")

    if not out:
        out = ["OTHER"]

    return out


def build_facility_types(row):
    # ✅ ทุกศูนย์มีพื้นฐานเสมอ
    out = ["ELECTRICITY", "WATER"]

    # ✅ ถ้า restroomCount > 0 ให้เพิ่ม RESTROOM
    rc = row.get("restroomCount")
    if isinstance(rc, (int, float)) and rc > 0:
        out.append("RESTROOM")

    if row.get("fac_t"):
        out.append("TENT")
    if row.get("fac_u"):
        out.append("PET")
    if row.get("fac_v"):
        out.append("KITCHEN")
    if row.get("fac_aa"):
        out.append("INTERNET")
    if row.get("fac_ab"):
        out.append("SECURITY")

    out = list(dict.fromkeys(out))
    return out


# ------------------------------
# DB ops
# ------------------------------
def quote_cols(payload: dict) -> str:
    return ", ".join([f'"{k}"' if any(c.isupper() for c in k) else k for k in payload.keys()])


def insert_shelter(cur, shelter_table, data, region_value: str | None):
    now = datetime.now(TZ)
    shelter_id = str(uuid.uuid4())

    payload = {
        "id": shelter_id,
        "code": None,
        "name": data.get("name", ""),  # ✅ อนุญาตว่าง
        "description": None,
        "shelterType": data["shelterType"],
        "responsibleAgency": data.get("responsibleAgency") or "-",
        "capacity": data["capacity"],
        "coordinatorName": data.get("coordinatorName"),
        "coordinatorPhone": data.get("coordinatorPhone"),
        "contactEmail": None,
        "status": "ACTIVE",
        "isDraft": False,
        "isOpen": False,
        "autoAccept": True,
        "isPublic": True,

        # ✅ ใส่ addressLine จาก Excel (แทน village)
        "addressLine": data.get("addressLine"),
        "village": None,

        "province": data.get("province"),
        "district": data.get("district"),
        "subdistrict": data.get("subdistrict"),
        "region": region_value,
        "postcode": data.get("postcode"),
        "latitude": data.get("latitude"),
        "longitude": data.get("longitude"),
        "landmark": None,
        "restroomCount": data.get("restroomCount") if data.get("restroomCount") is not None else 0,
        "parkingCount": data.get("parkingCount") if data.get("parkingCount") is not None else 0,
        "otherFacilitiesNote": None,
        "currentResidents": 0,
        "vulnerableResidents": 0,
        "occupancyUpdatedAt": now,
        "publishedAt": now,
        "deletedAt": None,
        "createdAt": now,
        "updatedAt": now,
    }

    cols = quote_cols(payload)
    ph = ", ".join(["%s"] * len(payload))
    sql = f"INSERT INTO {shelter_table} ({cols}) VALUES ({ph});"
    cur.execute(sql, tuple(payload.values()))
    return shelter_id


def insert_disasters(cur, disasters_table, shelter_id: str, types: list[str]):
    inserted = 0
    sid = str(shelter_id)

    for t in types:
        did = str(uuid.uuid4())
        sql = f"""
            INSERT INTO {disasters_table} (id, type, "shelterId")
            SELECT %s, %s, %s
            WHERE NOT EXISTS (
                SELECT 1 FROM {disasters_table}
                WHERE type = %s
                  AND "shelterId"::text = %s
            );
        """
        cur.execute(sql, (did, t, sid, t, sid))
        inserted += cur.rowcount

    return inserted


def insert_facilities(cur, facility_table, shelter_id: str, types: list[str]):
    inserted = 0
    sid = str(shelter_id)

    for t in types:
        fid = str(uuid.uuid4())
        sql = f"""
            INSERT INTO {facility_table} (id, type, "shelterId")
            SELECT %s, %s, %s
            WHERE NOT EXISTS (
                SELECT 1 FROM {facility_table}
                WHERE type = %s
                  AND "shelterId"::text = %s
            );
        """
        cur.execute(sql, (fid, t, sid, t, sid))
        inserted += cur.rowcount

    return inserted


def count_shelters_in_province(cur, shelter_table, province: str) -> int:
    cur.execute(f"SELECT COUNT(*) AS c FROM {shelter_table} WHERE province = %s;", (province,))
    return int(cur.fetchone()["c"])


def purge_province(cur, province: str, shelter_table, disasters_table, facility_table, dailystats_table, resident_table):
    subquery = f"(SELECT id::text FROM {shelter_table} WHERE province = %s)"

    cur.execute(f'DELETE FROM {disasters_table} WHERE "shelterId"::text IN {subquery};', (province,))
    del_dis = cur.rowcount

    cur.execute(f'DELETE FROM {facility_table} WHERE "shelterId"::text IN {subquery};', (province,))
    del_fac = cur.rowcount

    cur.execute(f'DELETE FROM {dailystats_table} WHERE "shelterId"::text IN {subquery};', (province,))
    del_stats = cur.rowcount

    cur.execute(f'DELETE FROM {resident_table} WHERE "shelterId"::text IN {subquery};', (province,))
    del_res = cur.rowcount

    cur.execute(f"DELETE FROM {shelter_table} WHERE province = %s;", (province,))
    del_shel = cur.rowcount

    return {
        "deleted_disasters": del_dis,
        "deleted_facilities": del_fac,
        "deleted_daily_stats": del_stats,
        "deleted_residents": del_res,
        "deleted_shelters": del_shel,
    }


# ------------------------------
# main
# ------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True, help="path to .xlsx")
    ap.add_argument("--sheet", default=None, help="sheet name (optional)")
    ap.add_argument("--start-row", type=int, default=2, help="start row (default 2)")
    ap.add_argument("--province", default=None, help="import as group by this province (จะลบจังหวัดนี้ก่อน import)")
    ap.add_argument("--region", default=None, help="กำหนด region ให้จังหวัดนี้ (ใช้คู่กับ --province)")
    ap.add_argument("--limit", type=int, default=None, help="limit after filtering")
    ap.add_argument("--dry-run", action="store_true", help="ไม่เขียน DB แค่พิมพ์สิ่งที่จะทำ")
    args = ap.parse_args()

    shelter_table = qname(T_SHELTER)
    disasters_table = qname(T_DISASTERS)
    facility_table = qname(T_FACILITY)
    dailystats_table = qname(T_DAILYSTATS)
    resident_table = qname(T_RESIDENT)

    region_default = os.getenv("REGION_DEFAULT")
    if args.province:
        if not args.region:
            raise ValueError("ต้องระบุ --region เมื่อใช้ --province (โหมดนำเข้าตามจังหวัด)")
        region_to_use = args.region
    else:
        region_to_use = args.region or region_default

    if not region_to_use:
        print("⚠️ region จะเป็น NULL (อาจ error ถ้า column region เป็น NOT NULL)")

    rows = read_excel_rows(args.excel, sheet_name=args.sheet, start_row=args.start_row)

    # ✅ โหมดกลุ่ม:
    # - คัดเฉพาะแถวที่ province ใน Excel = args.province หรือ province ว่าง
    # - แล้วบังคับ province ในข้อมูลที่จะ insert เป็น args.province
    if args.province:
        rows = [r for r in rows if (r.get("province") or "") in ("", args.province)]
        for r in rows:
            r["province"] = args.province

    if args.limit is not None:
        rows = rows[: args.limit]

    print(f"อ่านได้ {len(rows)} แถว (หลัง filter แล้ว)")

    inserted_shelters = 0
    inserted_disasters = 0
    inserted_facilities = 0
    failed_rows = 0

    with get_conn() as conn:
        try:
            with conn.cursor() as cur:
                # purge จังหวัดก่อน
                if args.province:
                    existing_cnt = count_shelters_in_province(cur, shelter_table, args.province)
                    print(f"\n[GROUP MODE] province='{args.province}' region='{region_to_use}'")
                    print(f"จะลบ Shelter เดิมในจังหวัดนี้ทั้งหมด = {existing_cnt} รายการ")

                    if args.dry_run:
                        print("(dry-run) จะลบ disasters/facilities/dailyStats/resident ที่ผูกกับ Shelter ของจังหวัดนี้ด้วย")
                    else:
                        summary = purge_province(
                            cur,
                            province=args.province,
                            shelter_table=shelter_table,
                            disasters_table=disasters_table,
                            facility_table=facility_table,
                            dailystats_table=dailystats_table,
                            resident_table=resident_table,
                        )
                        print("✅ purge done:", summary)

                # ✅ INSERT ใหม่ทุกแถว (ไม่สน name ซ้ำ/ว่าง)
                for i, r in enumerate(rows, start=1):
                    dis_types = build_disaster_types(r)
                    fac_types = build_facility_types(r)

                    if args.dry_run:
                        print(f"\n[{i}] name='{r.get('name','')}' (FORCE NEW)")
                        print("  disasters:", dis_types)
                        print("  facilities:", fac_types)
                        continue

                    # ✅ แถวไหนพัง ให้ข้ามแถวนั้น แต่ไม่ล้มทั้งงาน
                    cur.execute("SAVEPOINT sp_row;")
                    try:
                        shelter_id = insert_shelter(cur, shelter_table, r, region_to_use)
                        inserted_shelters += 1

                        inserted_disasters += insert_disasters(cur, disasters_table, shelter_id, dis_types)
                        inserted_facilities += insert_facilities(cur, facility_table, shelter_id, fac_types)

                        cur.execute("RELEASE SAVEPOINT sp_row;")
                    except Exception as e:
                        failed_rows += 1
                        cur.execute("ROLLBACK TO SAVEPOINT sp_row;")
                        cur.execute("RELEASE SAVEPOINT sp_row;")
                        print(f"❌ row {i} failed (skip): name='{r.get('name','')}' | err={type(e).__name__}: {e}")

                if not args.dry_run:
                    conn.commit()

        except Exception:
            conn.rollback()
            raise

    print("\n===== SUMMARY =====")
    print("inserted_shelters   =", inserted_shelters)
    if args.dry_run:
        print("(dry-run) ไม่ได้เขียน DB")
    else:
        print("inserted_disasters  =", inserted_disasters)
        print("inserted_facilities =", inserted_facilities)
        print("failed_rows (skip)  =", failed_rows)


if __name__ == "__main__":
    main()