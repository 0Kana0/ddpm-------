# import_local_users.py
import os
import re
import uuid
import argparse
from datetime import datetime
from zoneinfo import ZoneInfo

import psycopg
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

try:
    import bcrypt
except ImportError:
    raise SystemExit("Missing dependency: bcrypt  (pip install bcrypt)")

TZ = ZoneInfo("Asia/Bangkok")

KEYWORDS = ["อบต.", "เทศบาลตำบล"]


def now_th() -> datetime:
    return datetime.now(TZ)


def get_conninfo():
    host = os.getenv("PGHOST")
    port = os.getenv("PGPORT", "5432")
    dbname = os.getenv("PGDATABASE")
    user = os.getenv("PGUSER")
    password = os.getenv("PGPASSWORD")
    sslmode = os.getenv("PGSSL")

    if not all([host, dbname, user, password]):
        raise SystemExit("Missing env. Required: PGHOST PGDATABASE PGUSER PGPASSWORD (PGPORT optional)")

    conninfo = {
        "host": host,
        "port": int(port),
        "dbname": dbname,
        "user": user,
        "password": password,
    }
    if sslmode:
        conninfo["sslmode"] = sslmode
    return conninfo


def norm(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None


def digits_only(v) -> str | None:
    s = norm(v)
    if not s:
        return None
    d = re.sub(r"\D+", "", s)
    return d if d else None


def contains_keyword(localgov_raw: str | None) -> bool:
    if not localgov_raw:
        return False
    return any(k in localgov_raw for k in KEYWORDS)


def clean_tambon_from_col_c(localgov_raw: str | None) -> str | None:
    """
    tambon = column C แต่ตัดคำว่า 'อบต.' และ 'เทศบาลตำบล' ออก
    """
    s = norm(localgov_raw)
    if not s:
        return None

    # remove keywords
    for k in KEYWORDS:
        s = s.replace(k, "")

    # tidy spaces/punctuation
    s = s.replace(":", " ")
    s = re.sub(r"\s+", " ", s).strip()

    # remove leading punctuation/spaces
    s = s.lstrip(" .-–—\t")

    return s if s else None


def bcrypt_hash_password(plain: str, rounds: int = 12) -> str:
    salt = bcrypt.gensalt(rounds=rounds)
    hashed = bcrypt.hashpw(plain.encode("utf-8"), salt)
    return hashed.decode("utf-8")


def is_header_row(v) -> bool:
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in {"username", "user", "ชื่อผู้ใช้"}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True, help="path to .xlsx")
    ap.add_argument("--sheet", default=None, help="sheet name (default active sheet)")
    ap.add_argument("--table", default='public."User"', help='target table, e.g. public."User"')
    ap.add_argument("--password", default="12345678", help="plain password to hash (default 12345678)")
    ap.add_argument("--bcrypt-rounds", type=int, default=12, help="bcrypt cost (default 12)")
    ap.add_argument("--dry-run", action="store_true", help="print actions only, no DB write")
    ap.add_argument("--upsert", action="store_true", help="ON CONFLICT(username) DO UPDATE")
    args = ap.parse_args()

    wb = load_workbook(args.excel)
    ws = wb[args.sheet] if args.sheet else wb.active

    start_row = 2 if is_header_row(ws["B1"].value) else 1  # header มักอยู่แถว 1

    conninfo = get_conninfo()

    # ✅ เพิ่ม tambon และให้ localGov เป็น raw
    insert_sql = f"""
    INSERT INTO {args.table} (
      id, username, "passwordHash",
      role, province, amphoe, tambon, "localGov",
      "isActive", "roleConfigId",
      email, "emailVerified",
      "twoFactorEnabled", "failedLoginAttempts",
      "readOnly", "createdAt", "updatedAt"
    )
    VALUES (
      %(id)s, %(username)s, %(passwordHash)s,
      %(role)s, %(province)s, %(amphoe)s, %(tambon)s, %(localGov)s,
      %(isActive)s, %(roleConfigId)s,
      %(email)s, %(emailVerified)s,
      %(twoFactorEnabled)s, %(failedLoginAttempts)s,
      %(readOnly)s, %(createdAt)s, %(updatedAt)s
    )
    """

    if args.upsert:
        insert_sql += """
        ON CONFLICT (username) DO UPDATE SET
          "passwordHash" = EXCLUDED."passwordHash",
          role = EXCLUDED.role,
          province = EXCLUDED.province,
          amphoe = EXCLUDED.amphoe,
          tambon = EXCLUDED.tambon,
          "localGov" = EXCLUDED."localGov",
          "isActive" = EXCLUDED."isActive",
          "roleConfigId" = EXCLUDED."roleConfigId",
          email = EXCLUDED.email,
          "emailVerified" = EXCLUDED."emailVerified",
          "updatedAt" = EXCLUDED."updatedAt"
        """

    rows = []
    skipped_no_keyword = 0
    skipped_no_username = 0

    for r in range(start_row, ws.max_row + 1):
        # ตาม flow ใหม่:
        # B=username, C=localGov(raw) + tambon(cleaned), D=amphoe, E=province
        username_raw = ws[f"B{r}"].value
        col_c_raw = norm(ws[f"C{r}"].value)     # raw -> localGov
        amphoe = norm(ws[f"D{r}"].value)
        province = norm(ws[f"E{r}"].value)

        # บันทึกเฉพาะแถวที่ col C มี keyword
        if not contains_keyword(col_c_raw):
            skipped_no_keyword += 1
            continue

        # username เอาเฉพาะตัวเลข
        username = digits_only(username_raw)
        if not username:
            skipped_no_username += 1
            continue

        payload = {
            "id": str(uuid.uuid4()),
            "username": username,
            "passwordHash": bcrypt_hash_password(args.password, rounds=args.bcrypt_rounds),
            "role": "LOCAL",
            "province": province,
            "amphoe": amphoe,
            # ✅ tambon = col C cleaned (ตัด อบต./เทศบาลตำบล)
            "tambon": clean_tambon_from_col_c(col_c_raw),
            # ✅ localGov = col C raw
            "localGov": col_c_raw,
            "isActive": True,
            "roleConfigId": "LOCAL",
            # ตั้งเป็น NULL เพื่อไม่ชน UNIQUE email
            "email": None,
            "emailVerified": False,
            "twoFactorEnabled": False,
            "failedLoginAttempts": 0,
            "readOnly": False,
            "createdAt": now_th(),
            "updatedAt": now_th(),
        }
        rows.append(payload)

    print(f"FOUND rows to import: {len(rows)}")
    print(f"SKIP (no keyword in col C): {skipped_no_keyword}")
    print(f"SKIP (username not numeric/empty): {skipped_no_username}")

    if args.dry_run:
        for p in rows[:15]:
            print(
                f"- username={p['username']} province={p['province']} amphoe={p['amphoe']} "
                f"tambon={p['tambon']} localGov(raw)={p['localGov']}"
            )
        print("DRY RUN: no DB changes.")
        return

    with psycopg.connect(**conninfo) as conn:
        with conn.cursor() as cur:
            for p in rows:
                cur.execute(insert_sql, p)
        conn.commit()

    print("DONE ✅ Imported LOCAL users successfully.")


if __name__ == "__main__":
    main()