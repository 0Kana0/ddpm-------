# export_users_excel.py
import os
import argparse
from datetime import datetime
from zoneinfo import ZoneInfo

import psycopg
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

TZ = ZoneInfo("Asia/Bangkok")


def get_conninfo():
    # env: PGHOST PGPORT PGDATABASE PGUSER PGPASSWORD PGSSL(optional)
    host = os.getenv("PGHOST")
    port = os.getenv("PGPORT", "5432")
    dbname = os.getenv("PGDATABASE")
    user = os.getenv("PGUSER")
    password = os.getenv("PGPASSWORD")
    sslmode = os.getenv("PGSSL")

    if not all([host, dbname, user, password]):
        raise SystemExit(
            "Missing env. Required: PGHOST PGDATABASE PGUSER PGPASSWORD (PGPORT optional)"
        )

    conninfo = {
        "host": host,
        "port": int(port),
        "dbname": dbname,
        "user": user,
        "password": password,
    }
    if sslmode:
        conninfo["sslmode"] = sslmode
    return conninfo


def autosize_columns_with_padding(ws, padding=4, min_width=12, max_width=60):
    """
    ปรับความกว้างคอลัมน์ตามข้อมูลจริง + เผื่อ padding
    - padding: เพิ่มความกว้างเผื่ออ่านง่าย
    - min_width/max_width: กันแคบ/กว้างเกินไป
    """
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))

        width = max_len + padding
        width = max(min_width, width)
        width = min(max_width, width)

        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--table", default='public."User"', help='e.g. public."User" or public.users')
    ap.add_argument("--out", default=None, help="output xlsx path")
    ap.add_argument("--sheet", default="users", help="sheet name")
    ap.add_argument("--only-active", action="store_true", help='filter WHERE "isActive"=true')
    ap.add_argument("--include-deleted", action="store_true", help='include deleted rows (ignore deletedAt filter)')
    ap.add_argument("--padding", type=int, default=4, help="autosize padding (default 4)")
    ap.add_argument("--min-width", type=int, default=12, help="min column width (default 12)")
    ap.add_argument("--max-width", type=int, default=60, help="max column width (default 60)")
    args = ap.parse_args()

    # ชื่อไฟล์ default
    out_path = args.out or "รายการ_users.xlsx"

    where = []
    if not args.include_deleted:
        where.append('"deletedAt" IS NULL')
    if args.only_active:
        where.append('"isActive" = true')

    # ตัด CENTRAL, SUPERADMIN ออก
    where.append("role NOT IN ('CENTRAL', 'SUPERADMIN')")

    where_sql = "WHERE " + " AND ".join(where) if where else ""

    query = f"""
    SELECT
      username,
      role,
      province,
      amphoe,
      "localGov"
    FROM {args.table}
    {where_sql}
    ORDER BY username
    """

    conninfo = get_conninfo()

    with psycopg.connect(**conninfo) as conn:
        with conn.cursor() as cur:
            cur.execute(query)
            rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = args.sheet

    headers = ["username", "password", "role", "province", "amphoe", "localGov"]
    ws.append(headers)

    # header style
    header_font = Font(bold=True)
    header_align = Alignment(vertical="center")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = header_align

    for (username, role, province, amphoe, localGov) in rows:
        ws.append([
            username,
            "12345678",  # password fix ทุกแถว
            role,
            province,
            amphoe,
            localGov,
        ])

    # autosize + padding
    autosize_columns_with_padding(
        ws,
        padding=args.padding,
        min_width=args.min_width,
        max_width=args.max_width,
    )

    wb.save(out_path)
    print(f"DONE ✅ Exported {len(rows)} rows -> {out_path}")


if __name__ == "__main__":
    main()