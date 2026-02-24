# import_users.py
import os
import uuid
import argparse
from datetime import datetime
from zoneinfo import ZoneInfo

import psycopg
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

try:
    import bcrypt
except ImportError:
    raise SystemExit("Missing dependency: bcrypt  (pip install bcrypt)")

TZ = ZoneInfo("Asia/Bangkok")


def now_th() -> datetime:
    return datetime.now(TZ)


def get_conninfo():
    # ใช้ env: PGHOST PGPORT PGDATABASE PGUSER PGPASSWORD PGSSL(optional)
    host = os.getenv("PGHOST")
    port = os.getenv("PGPORT", "5432")
    dbname = os.getenv("PGDATABASE")
    user = os.getenv("PGUSER")
    password = os.getenv("PGPASSWORD")
    sslmode = os.getenv("PGSSL")  # require/disable/verify-full...

    if not all([host, dbname, user, password]):
        raise SystemExit("Missing env. Required: PGHOST PGDATABASE PGUSER PGPASSWORD (PGPORT optional)")

    conninfo = {
        "host": host,
        "port": int(port),
        "dbname": dbname,
        "user": user,
        "password": password,
    }
    if sslmode:
        conninfo["sslmode"] = sslmode
    return conninfo


def bcrypt_hash_password(plain: str, rounds: int = 12) -> str:
    # ให้ prefix ออกมาเป็น $2b$... เหมือนตัวอย่าง
    salt = bcrypt.gensalt(rounds=rounds)
    hashed = bcrypt.hashpw(plain.encode("utf-8"), salt)
    return hashed.decode("utf-8")


def is_header_row(a1_value) -> bool:
    if a1_value is None:
        return False
    s = str(a1_value).strip().lower()
    return s in {"username", "user", "ชื่อผู้ใช้"}


def norm_cell(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True, help="path to .xlsx")
    ap.add_argument("--sheet", default=None, help="sheet name (default active sheet)")
    ap.add_argument("--table", default='public."User"', help='target table, e.g. public."User"')
    ap.add_argument("--email-domain", default="example.invalid", help="email domain for generated email")
    ap.add_argument("--password", default="12345678", help="plain password to hash (default 12345678)")
    ap.add_argument("--bcrypt-rounds", type=int, default=12, help="bcrypt cost (default 12)")
    ap.add_argument("--dry-run", action="store_true", help="print actions only, no DB write")
    ap.add_argument("--upsert", action="store_true", help="ON CONFLICT(username) DO UPDATE")
    args = ap.parse_args()

    wb = load_workbook(args.excel)
    ws = wb[args.sheet] if args.sheet else wb.active

    start_row = 2 if is_header_row(ws["A1"].value) else 1

    # hash ครั้งเดียว เพราะรหัสเดียวกันทุก user (แต่ salt จะต่างถ้า hash ใหม่ทุกแถว)
    # ถ้าต้องการให้ทุกคน hash ต่างกัน แนะนำให้ย้ายไป hash ใน loop
    # ที่นี่เลือก "hash แยกต่อแถว" เพื่อให้ปลอดภัยกว่า
    conninfo = get_conninfo()

    insert_sql = f"""
    INSERT INTO {args.table} (
      id, username, "passwordHash",
      role, province, amphoe,
      "isActive", "roleConfigId",
      email, "emailVerified",
      "twoFactorEnabled", "failedLoginAttempts",
      "readOnly", "createdAt", "updatedAt"
    )
    VALUES (
      %(id)s, %(username)s, %(passwordHash)s,
      %(role)s, %(province)s, %(amphoe)s,
      %(isActive)s, %(roleConfigId)s,
      %(email)s, %(emailVerified)s,
      %(twoFactorEnabled)s, %(failedLoginAttempts)s,
      %(readOnly)s, %(createdAt)s, %(updatedAt)s
    )
    """

    if args.upsert:
        insert_sql += """
        ON CONFLICT (username) DO UPDATE SET
          "passwordHash" = EXCLUDED."passwordHash",
          role = EXCLUDED.role,
          province = EXCLUDED.province,
          amphoe = EXCLUDED.amphoe,
          "isActive" = EXCLUDED."isActive",
          "roleConfigId" = EXCLUDED."roleConfigId",
          email = EXCLUDED.email,
          "emailVerified" = EXCLUDED."emailVerified",
          "updatedAt" = EXCLUDED."updatedAt"
        """

    rows = []
    for r in range(start_row, ws.max_row + 1):
        username = norm_cell(ws[f"A{r}"].value)
        role = norm_cell(ws[f"B{r}"].value)
        province = norm_cell(ws[f"C{r}"].value)
        amphoe = norm_cell(ws[f"D{r}"].value)

        if not username:
            continue

        # hash แยกต่อแถว (salt ต่างกัน)
        pwd_hash = bcrypt_hash_password(args.password, rounds=args.bcrypt_rounds)

        payload = {
            "id": str(uuid.uuid4()),
            "username": username,
            "passwordHash": pwd_hash,
            "role": role,
            "province": province,
            "amphoe": amphoe,
            "isActive": True,
            "roleConfigId": role,  # ตามโจทย์: column B
            "email": None,
            "emailVerified": False,
            "twoFactorEnabled": False,
            "failedLoginAttempts": 0,
            "readOnly": False,
            "createdAt": now_th(),
            "updatedAt": now_th(),
        }
        rows.append(payload)

    print(f"FOUND rows to import: {len(rows)} (from Excel {args.excel})")
    if args.dry_run:
        for p in rows[:10]:
            print(f"- username={p['username']} role={p['role']} province={p['province']} amphoe={p['amphoe']} email={p['email']}")
        print("DRY RUN: no DB changes.")
        return

    with psycopg.connect(**conninfo) as conn:
        with conn.cursor() as cur:
            for p in rows:
                cur.execute(insert_sql, p)
        conn.commit()

    print("DONE ✅ Imported users successfully.")


if __name__ == "__main__":
    main()