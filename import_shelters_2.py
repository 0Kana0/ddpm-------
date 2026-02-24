import os
import re
import uuid
import argparse
from datetime import datetime
from zoneinfo import ZoneInfo

from dotenv import load_dotenv
import psycopg
from psycopg.rows import dict_row
from openpyxl import load_workbook

load_dotenv()

# ===== Table names =====
T_SHELTER = "Shelter"
T_DISASTERS = "disasters"
T_FACILITY = "ShelterFacility"
T_DAILYSTATS = "ShelterDailyStats"
T_RESIDENT = "Resident"

TZ = ZoneInfo("Asia/Bangkok")


def qname(name: str) -> str:
    return f'"{name}"' if any(c.isupper() for c in name) else name


def get_conn():
    host = os.getenv("PGHOST", "127.0.0.1")
    port = int(os.getenv("PGPORT", "5432"))
    dbname = os.getenv("PGDATABASE")
    user = os.getenv("PGUSER")
    password = os.getenv("PGPASSWORD")
    sslmode = os.getenv("PGSSL", "require")

    if not all([dbname, user, password]):
        raise ValueError("Missing env: PGDATABASE / PGUSER / PGPASSWORD")

    return psycopg.connect(
        host=host,
        port=port,
        dbname=dbname,
        user=user,
        password=password,
        sslmode=sslmode,
        connect_timeout=15,
        autocommit=False,
        row_factory=dict_row,
    )


# ------------------------------
# helpers
# ------------------------------
def norm_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None


def is_has(v) -> bool:
    """
    True เมื่อเจอ:
      - "มี"
      - หรือมีคำว่า "รองรับ" / "รับรอง"
    False เมื่อเจอ:
      - "ไม่มี" / "ไม่รองรับ" / "ไม่รับรอง" / หรือขึ้นต้นด้วย "ไม่"
    """
    s = norm_text(v)
    if not s:
        return False

    s = s.strip()

    # เคสปฏิเสธ
    if "ไม่มี" in s or "ไม่รองรับ" in s or "ไม่รับรอง" in s or s.startswith("ไม่"):
        return False

    # เคสยืนยัน
    if s == "มี":
        return True
    if "มี" in s:
        return True
    if "รองรับ" in s:
        return True
    if "รับรอง" in s:
        return True

    return False


def to_int(v, default=None):
    s = norm_text(v)
    if s is None:
        return default
    try:
        return int(float(s))
    except Exception:
        return default


def to_float(v, default=None):
    s = norm_text(v)
    if s is None:
        return default
    try:
        return float(s)
    except Exception:
        return default


def clean_phone(v):
    s = norm_text(v)
    if not s:
        return None
    s = s.replace(" ", "").replace("-", "")
    s = re.sub(r"[^0-9+]", "", s)
    return s or None


def map_shelter_type(v):
    """
    กติกาใหม่:
      - ถ้าเจอคำว่า "ศูนย์พักพิงพระราชทาน" (หรือค่าเป็น ROYAL) -> ROYAL
      - นอกนั้นทั้งหมด -> SYSTEM
    """
    s = norm_text(v)
    if not s:
        return "SYSTEM"

    s_upper = s.upper()
    if s_upper == "ROYAL":
        return "ROYAL"

    if "ศูนย์พักพิงพระราชทาน" in s:
        return "ROYAL"

    return "SYSTEM"


# ------------------------------
# Excel reader (mapping ใหม่)
# ------------------------------
def read_excel_rows(excel_path, sheet_name=None, start_row=2):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = []
    for r in range(start_row, ws.max_row + 1):
        # ✅ กันแถวว่างท้ายไฟล์ (แต่ "name ว่าง" ยังอนุญาต ถ้าแถวมีข้อมูลอย่างอื่น)
        check_cells = [
            ws[f"C{r}"].value,  # name
            ws[f"D{r}"].value,  # shelterType
            ws[f"E{r}"].value, ws[f"F{r}"].value,
            ws[f"H{r}"].value, ws[f"J{r}"].value, ws[f"L{r}"].value,
            ws[f"O{r}"].value, ws[f"P{r}"].value,
            ws[f"R{r}"].value, ws[f"S{r}"].value,
            ws[f"W{r}"].value, ws[f"X{r}"].value,
            ws[f"AE{r}"].value, ws[f"AF{r}"].value,
            ws[f"AD{r}"].value, ws[f"AG{r}"].value, ws[f"AH{r}"].value,
            ws[f"AI{r}"].value, ws[f"AJ{r}"].value, ws[f"AK{r}"].value,
            ws[f"AL{r}"].value, ws[f"AM{r}"].value, ws[f"AN{r}"].value, ws[f"AO{r}"].value,
        ]
        if all(norm_text(v) is None for v in check_cells):
            continue

        # column C -> name (✅ อนุญาตว่าง)
        name = norm_text(ws[f"C{r}"].value)
        if name is None:
            name = ""  # ไม่สน name ว่าง ก็ยัง insert

        row = {
            # Shelter
            "name": name,
            "shelterType": map_shelter_type(ws[f"D{r}"].value),     # column D
            "responsibleAgency": norm_text(ws[f"AF{r}"].value),     # column AF
            "capacity": to_int(ws[f"R{r}"].value, default=None),    # column R

            "coordinatorName": norm_text(ws[f"W{r}"].value),        # column W
            "coordinatorPhone": clean_phone(ws[f"X{r}"].value),     # column X

            "addressLine": norm_text(ws[f"E{r}"].value),            # column E (เปลี่ยนจาก village → addressLine)
            "postcode": norm_text(ws[f"F{r}"].value),               # column F
            "province": norm_text(ws[f"H{r}"].value),               # column H (แต่โหมด --province จะ override)
            "district": norm_text(ws[f"J{r}"].value),               # column J
            "subdistrict": norm_text(ws[f"L{r}"].value),            # column L

            "latitude": to_float(ws[f"O{r}"].value),                # column O
            "longitude": to_float(ws[f"P{r}"].value),               # column P

            "restroomCount": to_int(ws[f"S{r}"].value, default=0),  # column S
            "parkingCount": to_int(ws[f"AE{r}"].value, default=0),  # column AE

            # disasters flags: AL,AM,AN,AO
            "dis_al": is_has(ws[f"AL{r}"].value),  # FLOOD
            "dis_am": is_has(ws[f"AM{r}"].value),  # STORM
            "dis_an": is_has(ws[f"AN{r}"].value),  # FIRE
            "dis_ao": is_has(ws[f"AO{r}"].value),  # EARTHQUAKE

            # facilities flags: AD,AG,AH,AI,AJ,AK
            "fac_ad": is_has(ws[f"AD{r}"].value),  # ELECTRICITY
            "fac_ag": is_has(ws[f"AG{r}"].value),  # INTERNET
            "fac_ah": is_has(ws[f"AH{r}"].value),  # SECURITY
            "fac_ai": is_has(ws[f"AI{r}"].value),  # TENT
            "fac_aj": is_has(ws[f"AJ{r}"].value),  # PET
            "fac_ak": is_has(ws[f"AK{r}"].value),  # KITCHEN
        }
        rows.append(row)

    return rows


def build_disaster_types(row):
    out = []
    if row.get("dis_al"):
        out.append("FLOOD")
    if row.get("dis_am"):
        out.append("STORM")
    if row.get("dis_an"):
        out.append("FIRE")
    if row.get("dis_ao"):
        out.append("EARTHQUAKE")

    if not out:
        out = ["OTHER"]

    return out


def build_facility_types(row):
    # ✅ ทุกศูนย์มีพื้นฐานเสมอ
    out = ["ELECTRICITY", "WATER"]

    # ✅ เพิ่ม RESTROOM เมื่อมีห้องน้ำ > 0
    rc = row.get("restroomCount") or 0
    if isinstance(rc, (int, float)) and rc > 0:
        out.append("RESTROOM")

    # ✅ เพิ่มตามคอลัมน์ (มี/รองรับ/รับรอง)
    if row.get("fac_ad"):
        out.append("ELECTRICITY")  # มีพื้นฐานแล้ว แต่ไม่เป็นไร
    if row.get("fac_ai"):
        out.append("TENT")
    if row.get("fac_aj"):
        out.append("PET")
    if row.get("fac_ak"):
        out.append("KITCHEN")
    if row.get("fac_ag"):
        out.append("INTERNET")
    if row.get("fac_ah"):
        out.append("SECURITY")

    out = list(dict.fromkeys(out))
    return out


# ------------------------------
# DB ops
# ------------------------------
def quote_cols(payload: dict) -> str:
    return ", ".join([f'"{k}"' if any(c.isupper() for c in k) else k for k in payload.keys()])


def insert_shelter(cur, shelter_table, data, region_value: str | None):
    now = datetime.now(TZ)
    shelter_id = str(uuid.uuid4())

    payload = {
        "id": shelter_id,
        "code": None,

        "name": data.get("name", ""),  # ✅ อนุญาตว่าง

        "description": None,
        "shelterType": data["shelterType"],
        "responsibleAgency": data.get("responsibleAgency") or "-",
        "capacity": data["capacity"],

        "coordinatorName": data.get("coordinatorName"),
        "coordinatorPhone": data.get("coordinatorPhone"),

        "contactEmail": None,

        "status": "ACTIVE",
        "isDraft": False,
        "isOpen": False,
        "autoAccept": True,
        "isPublic": True,

        "addressLine": data.get("addressLine"),

        "village": None,
        "province": data.get("province"),
        "district": data.get("district"),
        "subdistrict": data.get("subdistrict"),
        "region": region_value,
        "postcode": data.get("postcode"),

        "latitude": data.get("latitude"),
        "longitude": data.get("longitude"),

        "landmark": None,

        "restroomCount": data.get("restroomCount") or 0,
        "parkingCount": data.get("parkingCount") or 0,

        "otherFacilitiesNote": None,

        "currentResidents": 0,
        "vulnerableResidents": 0,

        "occupancyUpdatedAt": now,
        "publishedAt": now,
        "deletedAt": None,
        "createdAt": now,
        "updatedAt": now,
    }

    cols = quote_cols(payload)
    ph = ", ".join(["%s"] * len(payload))
    sql = f"INSERT INTO {shelter_table} ({cols}) VALUES ({ph});"

    try:
        cur.execute(sql, tuple(payload.values()))
        return shelter_id
    except psycopg.errors.UniqueViolation:
        # ✅ ถ้ามี UNIQUE constraint แล้วชน: เติม suffix ให้ name เพื่อให้ insert ได้ "ทุกแถว"
        cur.execute("ROLLBACK;")  # psycopg3 ต้อง rollback transaction state ของ statement ที่ fail
        cur.execute("BEGIN;")

        original = payload.get("name", "")
        suffix = str(uuid.uuid4())[:8]
        if original is None or str(original).strip() == "":
            payload["name"] = f"UNNAMED-{suffix}"
        else:
            payload["name"] = f"{original} ({suffix})"

        cols2 = quote_cols(payload)
        ph2 = ", ".join(["%s"] * len(payload))
        sql2 = f"INSERT INTO {shelter_table} ({cols2}) VALUES ({ph2});"
        cur.execute(sql2, tuple(payload.values()))
        print(f"⚠️ name ซ้ำ/ชน UNIQUE → เปลี่ยนชื่อชั่วคราวเป็น: {payload['name']}")
        return shelter_id
    except psycopg.errors.NotNullViolation:
        # ✅ ถ้า name/field อื่นโดน NOT NULL แต่เป็น None: เติมค่าขั้นต่ำ
        cur.execute("ROLLBACK;")
        cur.execute("BEGIN;")
        if payload.get("name") is None:
            payload["name"] = ""
        if payload.get("responsibleAgency") is None:
            payload["responsibleAgency"] = "-"
        cols2 = quote_cols(payload)
        ph2 = ", ".join(["%s"] * len(payload))
        sql2 = f"INSERT INTO {shelter_table} ({cols2}) VALUES ({ph2});"
        cur.execute(sql2, tuple(payload.values()))
        return shelter_id


def insert_disasters(cur, disasters_table, shelter_id: str, types: list[str]):
    inserted = 0
    sid = str(shelter_id)

    for t in types:
        did = str(uuid.uuid4())
        sql = f"""
            INSERT INTO {disasters_table} (id, type, "shelterId")
            SELECT %s, %s, %s
            WHERE NOT EXISTS (
                SELECT 1 FROM {disasters_table}
                WHERE type = %s
                  AND "shelterId"::text = %s
            );
        """
        cur.execute(sql, (did, t, sid, t, sid))
        inserted += cur.rowcount

    return inserted


def insert_facilities(cur, facility_table, shelter_id: str, types: list[str]):
    inserted = 0
    sid = str(shelter_id)

    for t in types:
        fid = str(uuid.uuid4())
        sql = f"""
            INSERT INTO {facility_table} (id, type, "shelterId")
            SELECT %s, %s, %s
            WHERE NOT EXISTS (
                SELECT 1 FROM {facility_table}
                WHERE type = %s
                  AND "shelterId"::text = %s
            );
        """
        cur.execute(sql, (fid, t, sid, t, sid))
        inserted += cur.rowcount

    return inserted


def count_shelters_in_province(cur, shelter_table, province: str) -> int:
    cur.execute(f"SELECT COUNT(*) AS c FROM {shelter_table} WHERE province = %s;", (province,))
    return int(cur.fetchone()["c"])


def purge_province(cur, province: str, shelter_table, disasters_table, facility_table, dailystats_table, resident_table):
    subquery = f"(SELECT id::text FROM {shelter_table} WHERE province = %s)"

    cur.execute(f'DELETE FROM {disasters_table} WHERE "shelterId"::text IN {subquery};', (province,))
    del_dis = cur.rowcount

    cur.execute(f'DELETE FROM {facility_table} WHERE "shelterId"::text IN {subquery};', (province,))
    del_fac = cur.rowcount

    cur.execute(f'DELETE FROM {dailystats_table} WHERE "shelterId"::text IN {subquery};', (province,))
    del_stats = cur.rowcount

    cur.execute(f'DELETE FROM {resident_table} WHERE "shelterId"::text IN {subquery};', (province,))
    del_res = cur.rowcount

    cur.execute(f"DELETE FROM {shelter_table} WHERE province = %s;", (province,))
    del_shel = cur.rowcount

    return {
        "deleted_disasters": del_dis,
        "deleted_facilities": del_fac,
        "deleted_daily_stats": del_stats,
        "deleted_residents": del_res,
        "deleted_shelters": del_shel,
    }


# ------------------------------
# main
# ------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True, help="path to .xlsx")
    ap.add_argument("--sheet", default=None, help="sheet name (optional)")
    ap.add_argument("--start-row", type=int, default=2, help="start row (default 2)")
    ap.add_argument("--province", default=None, help="import as group by this province (จะลบจังหวัดนี้ก่อน import)")
    ap.add_argument("--region", default=None, help="กำหนด region ให้จังหวัดนี้ (ใช้คู่กับ --province)")
    ap.add_argument("--limit", type=int, default=None, help="limit after filtering")
    ap.add_argument("--dry-run", action="store_true", help="ไม่เขียน DB แค่พิมพ์สิ่งที่จะทำ")
    args = ap.parse_args()

    shelter_table = qname(T_SHELTER)
    disasters_table = qname(T_DISASTERS)
    facility_table = qname(T_FACILITY)
    dailystats_table = qname(T_DAILYSTATS)
    resident_table = qname(T_RESIDENT)

    # region
    region_default = os.getenv("REGION_DEFAULT")
    if args.province:
        if not args.region:
            raise ValueError("ต้องระบุ --region เมื่อใช้ --province (โหมดนำเข้าตามจังหวัด)")
        region_to_use = args.region
    else:
        region_to_use = args.region or region_default

    if not region_to_use:
        print("⚠️ region จะเป็น NULL (อาจ error ถ้า column region เป็น NOT NULL)")

    # อ่าน excel
    rows = read_excel_rows(args.excel, sheet_name=args.sheet, start_row=args.start_row)

    # ✅ โหมดกลุ่ม: บังคับ province จาก args (ไม่สนค่าใน Excel)
    if args.province:
        for r in rows:
            r["province"] = args.province

    if args.limit is not None:
        rows = rows[: args.limit]

    print(f"อ่านได้ {len(rows)} แถว (หลัง filter แล้ว)")

    inserted_shelters = 0
    inserted_disasters = 0
    inserted_facilities = 0

    with get_conn() as conn:
        try:
            with conn.cursor() as cur:
                # purge จังหวัดก่อน (group mode)
                if args.province:
                    existing_cnt = count_shelters_in_province(cur, shelter_table, args.province)
                    print(f"\n[GROUP MODE] province='{args.province}' region='{region_to_use}'")
                    print(f"จะลบ Shelter เดิมในจังหวัดนี้ทั้งหมด = {existing_cnt} รายการ")

                    if args.dry_run:
                        print("(dry-run) จะลบ disasters/facilities/dailyStats/resident ที่ผูกกับ Shelter ของจังหวัดนี้ด้วย")
                    else:
                        summary = purge_province(
                            cur,
                            province=args.province,
                            shelter_table=shelter_table,
                            disasters_table=disasters_table,
                            facility_table=facility_table,
                            dailystats_table=dailystats_table,
                            resident_table=resident_table,
                        )
                        print("✅ purge done:", summary)

                # ✅ import แบบ "ไม่สนซ้ำ" = insert ใหม่ทุกแถว
                for i, r in enumerate(rows, start=1):
                    name = r.get("name", "")

                    dis_types = build_disaster_types(r)
                    fac_types = build_facility_types(r)

                    if args.dry_run:
                        print(f"\n[{i}] name='{name}'")
                        print("  shelter: NEW (force insert)")
                        print("  disasters:", dis_types)
                        print("  facilities:", fac_types)
                        shelter_id = "NEW_UUID"
                    else:
                        shelter_id = insert_shelter(cur, shelter_table, r, region_to_use)
                        inserted_shelters += 1
                        inserted_disasters += insert_disasters(cur, disasters_table, shelter_id, dis_types)
                        inserted_facilities += insert_facilities(cur, facility_table, shelter_id, fac_types)

                if not args.dry_run:
                    conn.commit()

        except Exception:
            conn.rollback()
            raise

    print("\n===== SUMMARY =====")
    print("inserted_shelters   =", inserted_shelters)
    if args.dry_run:
        print("(dry-run) ไม่ได้เขียน DB")
    else:
        print("inserted_disasters  =", inserted_disasters)
        print("inserted_facilities =", inserted_facilities)


if __name__ == "__main__":
    main()
